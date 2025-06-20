import streamlit as st
import openai
import tempfile
from openai import OpenAI
from openpyxl import load_workbook

client = OpenAI(api_key=st.secrets["OPENAI_API_KEY"])

def extract_summary_data(file):
    wb = load_workbook(file, data_only=True)

    # Convert all formulas to values
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    try:
                        cell.value = cell.value  # keep last calculated value
                        cell.data_type = "n"
                    except:
                        pass

    summary = {}
    for sheetname in wb.sheetnames:
        if sheetname.startswith("2025") or sheetname == "Γενικό Αποτέλεσμα":
            ws = wb[sheetname]
            try:
                summary[sheetname] = {
                    "Λογιστικό Αποτέλεσμα": ws["D5"].value,
                    "Σύνολο Εξόδων": ws["D6"].value,
                    "Σύνολο Εσόδων": ws["D7"].value,
                    "Προοδευτικό Αποτέλεσμα": ws["F5"].value,
                    "Προοδευτικά Έξοδα": ws["F6"].value,
                    "Προοδευτικά Έσοδα": ws["F7"].value,
                    "Πέρσι Έξοδα": ws["L6"].value,
                    "Πέρσι Έσοδα": ws["L7"].value,
                }
            except:
                continue
    return summary

def ask_gpt(question, context_data):
    context_text = "Τα δεδομένα αφορούν λογιστικά αποτελέσματα για το 2025 ανά μήνα.\n"
    for month, values in context_data.items():
        context_text += f"\nΜήνας: {month}\n"
        for key, val in values.items():
            context_text += f"- {key}: {val}\n"
    prompt = f"{context_text}\n\nΕρώτηση χρήστη:\n{question}\n\nΑπάντησε με σαφήνεια και σύντομα."

    response = client.chat.completions.create(
        model="gpt-4-1106-preview",
        messages=[
            {"role": "system", "content": "Είσαι ένας οικονομικός βοηθός που αναλύει δεδομένα από Excel."},
            {"role": "user", "content": prompt}
        ],
        temperature=0.4,
        max_tokens=500
    )
    return response.choices[0].message.content

# Streamlit UI
st.set_page_config(page_title="Βοηθός Excel με GPT-4", layout="centered")
st.title("📊 Βοηθός Excel με GPT-4")

uploaded_file = st.file_uploader("📁 Μεταφόρτωσε το τελικό αρχείο Excel", type=["xlsx"])

if uploaded_file:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(uploaded_file.read())
        tmp.flush()
        summary_data = extract_summary_data(tmp.name)

    st.success("✅ Τα δεδομένα φορτώθηκαν και οι τύποι μετατράπηκαν σε τιμές.")

    st.markdown("### ❓ Κάνε την ερώτησή σου")
    user_question = st.text_input("Ερώτηση (π.χ. Ποιος μήνας είχε τα περισσότερα έξοδα;)")

    if user_question:
        with st.spinner("🔍 Αξιολόγηση..."):
            gpt_response = ask_gpt(user_question, summary_data)
        st.markdown("### 🧠 Απάντηση GPT")
        st.markdown(gpt_response)
